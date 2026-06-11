"""
日报导出服务
支持 Excel 和 PDF 格式导出
"""
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from django.http import HttpResponse
from django.utils import timezone
from .services import snapshot_service


class ExportService:
    """导出服务"""

    STATUS_COLORS = {
        'NORMAL': '#00FF00',
        'PARTIAL': '#FFFF00',
        'ATTENTION': '#FFA500',
        'WARNING': '#FF0000',
    }

    STATUS_NAMES = {
        'NORMAL': '正常',
        'PARTIAL': '部分完成',
        'ATTENTION': '需关注',
        'WARNING': '异常',
    }

    @classmethod
    def export_daily_report_excel(cls, target_date, pen_id=None):
        """导出 Excel 格式的日报"""
        snapshot = snapshot_service.calculate_daily_snapshot(target_date, pen_id)

        wb = Workbook()
        ws = wb.active
        ws.title = f'巡检日报_{target_date}'

        header_font = Font(bold=True, size=14, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        sub_header_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells('A1:L1')
        ws['A1'] = f'养殖场巡检日报 - {target_date}'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 30

        avg_health_score = snapshot.get('average_health_score', 0)
        risk_distribution = snapshot.get('health_score_risk_distribution', {})
        risk_text = ', '.join([f'{k}:{v}' for k, v in risk_distribution.items()]) or '无数据'

        summary_data = [
            ['汇总信息', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['栏区总数', snapshot['total_pens'], '平均巡检完成率', f'{snapshot["average_inspection_completion_rate"]}%',
             '待处理异常事件', snapshot['total_open_incidents'], '今日异常总数', snapshot['total_abnormal_count'],
             '平均健康评分', avg_health_score, '风险分布', risk_text,
             '导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''],
        ]

        row = 3
        for row_data in summary_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = center_align
                cell.border = thin_border
                if col == 1:
                    cell.font = sub_header_font
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='栏区详情').font = sub_header_font
        row += 1

        detail_headers = [
            '栏区编号', '栏区名称', '养殖类型', '当前数量',
            '巡检完成率', '巡检次数', '异常指标数',
            '喂养次数', '总投喂量(kg)',
            '清洁次数', '总耗时(分钟)',
            '待处理事件', '今日新事件', '健康评分', '风险等级', '整体状态'
        ]

        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = sub_header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        row += 1

        for pen_snap in snapshot['pen_snapshots']:
            status_color = cls.STATUS_COLORS.get(pen_snap['overall_status'], '#FFFFFF')
            status_name = cls.STATUS_NAMES.get(pen_snap['overall_status'], '未知')
            health_score = pen_snap.get('health_score') or {}
            hs_score = health_score.get('total_score', '-')
            hs_risk = health_score.get('risk_level_display', '-')
            hs_color = health_score.get('risk_level') and {
                'EXCELLENT': '#00C853', 'GOOD': '#64DD17', 'NORMAL': '#FFD600',
                'WARNING': '#FF9100', 'DANGER': '#FF1744'
            }.get(health_score.get('risk_level'), '#FFFFFF') or '#FFFFFF'

            row_data = [
                pen_snap['pen_code'],
                pen_snap['pen_name'],
                pen_snap['livestock_type'],
                pen_snap['current_count'],
                f'{pen_snap["inspection"]["completion_rate"]}%',
                pen_snap['inspection']['inspection_count'],
                pen_snap['inspection']['abnormal_count'],
                pen_snap['feeding']['feeding_count'],
                pen_snap['feeding']['total_feed_amount'],
                pen_snap['cleaning']['cleaning_count'],
                pen_snap['cleaning']['total_duration_minutes'],
                pen_snap['incidents']['open_count'],
                pen_snap['incidents']['new_today_count'],
                hs_score,
                hs_risk,
                status_name,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = center_align if col != 2 else left_align
                cell.border = thin_border
                if col == 15:
                    cell.fill = PatternFill(start_color=hs_color.lstrip('#'), end_color=hs_color.lstrip('#'), fill_type='solid')
                if col == 16:
                    excel_color = status_color.lstrip('#')
                    cell.fill = PatternFill(start_color=excel_color, end_color=excel_color, fill_type='solid')
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='异常事件详情').font = sub_header_font
        row += 1

        incident_headers = [
            '栏区编号', '事件标题', '严重程度', '状态',
            '发生时间', '上报人', '处理状态', '处理人', '持续时间(小时)'
        ]

        for col, header in enumerate(incident_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = sub_header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
        row += 1

        for pen_snap in snapshot['pen_snapshots']:
            all_incidents = (pen_snap['incidents']['open_incidents'] +
                             pen_snap['incidents']['in_progress_incidents'] +
                             pen_snap['incidents']['new_today_incidents'])
            for incident in all_incidents:
                row_data = [
                    pen_snap['pen_code'],
                    incident['title'],
                    incident['severity_display'],
                    incident['status_display'],
                    incident['incident_time'],
                    incident['reporter_name'],
                    incident['status_display'],
                    '',
                    ''
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = center_align if col != 2 else left_align
                    cell.border = thin_border
                row += 1

        column_widths = [12, 20, 10, 10, 12, 10, 12, 10, 14, 10, 14, 12, 12, 10, 10, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="巡检日报_{target_date}.xlsx"'
        return response

    @classmethod
    def export_daily_report_pdf(cls, target_date, pen_id=None):
        """导出 PDF 格式的日报"""
        snapshot = snapshot_service.calculate_daily_snapshot(target_date, pen_id)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=20
        )
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            spaceBefore=10,
            spaceAfter=10
        )

        story = []

        story.append(Paragraph(f'养殖场巡检日报 - {target_date}', title_style))
        story.append(Spacer(1, 0.5 * cm))

        avg_health_score = snapshot.get('average_health_score', 0)
        summary_data = [
            ['栏区总数', str(snapshot['total_pens']),
             '平均巡检完成率', f'{snapshot["average_inspection_completion_rate"]}%',
             '平均健康评分', str(avg_health_score),
             '待处理异常', str(snapshot['total_open_incidents']),
             '今日异常总数', str(snapshot['total_abnormal_count'])],
            ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', '', '', '', '', '', '', ''],
        ]

        summary_table = Table(summary_data, colWidths=[2.5 * cm, 2 * cm, 3.5 * cm, 2 * cm, 3 * cm, 2 * cm, 3 * cm, 2 * cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E2F3')),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.8 * cm))

        story.append(Paragraph('栏区详情', section_style))

        detail_headers = [
            '栏区编号', '栏区名称', '养殖类型', '当前数量',
            '巡检完成率', '巡检次数', '异常指标',
            '喂养次数', '投喂量(kg)',
            '清洁次数', '耗时(分)',
            '待处理', '新事件', '评分', '风险', '状态'
        ]

        detail_data = [detail_headers]
        for pen_snap in snapshot['pen_snapshots']:
            status_name = cls.STATUS_NAMES.get(pen_snap['overall_status'], '未知')
            health_score = pen_snap.get('health_score') or {}
            hs_score = str(health_score.get('total_score', '-'))
            hs_risk = str(health_score.get('risk_level_display', '-'))
            detail_data.append([
                pen_snap['pen_code'],
                pen_snap['pen_name'],
                pen_snap['livestock_type'],
                str(pen_snap['current_count']),
                f'{pen_snap["inspection"]["completion_rate"]}%',
                str(pen_snap['inspection']['inspection_count']),
                str(pen_snap['inspection']['abnormal_count']),
                str(pen_snap['feeding']['feeding_count']),
                str(pen_snap['feeding']['total_feed_amount']),
                str(pen_snap['cleaning']['cleaning_count']),
                str(pen_snap['cleaning']['total_duration_minutes']),
                str(pen_snap['incidents']['open_count']),
                str(pen_snap['incidents']['new_today_count']),
                hs_score,
                hs_risk,
                status_name,
            ])

        detail_table = Table(detail_data, colWidths=[1.5 * cm] * 16)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        for i, pen_snap in enumerate(snapshot['pen_snapshots'], 1):
            status = pen_snap['overall_status']
            color = cls.STATUS_COLORS.get(status, '#FFFFFF')
            style_cmds.append(('BACKGROUND', (15, i), (15, i), colors.HexColor(color)))
            health_score = pen_snap.get('health_score') or {}
            risk_level = health_score.get('risk_level')
            if risk_level:
                hs_color = {
                    'EXCELLENT': '#00C853', 'GOOD': '#64DD17', 'NORMAL': '#FFD600',
                    'WARNING': '#FF9100', 'DANGER': '#FF1744'
                }.get(risk_level, '#FFFFFF')
                style_cmds.append(('BACKGROUND', (14, i), (14, i), colors.HexColor(hs_color)))

        detail_table.setStyle(TableStyle(style_cmds))
        story.append(detail_table)
        story.append(Spacer(1, 0.8 * cm))

        story.append(Paragraph('异常事件详情', section_style))

        incident_headers = [
            '栏区编号', '事件标题', '严重程度', '状态', '发生时间', '上报人'
        ]

        all_incidents = []
        for pen_snap in snapshot['pen_snapshots']:
            pen_incidents = (pen_snap['incidents']['open_incidents'] +
                             pen_snap['incidents']['in_progress_incidents'] +
                             pen_snap['incidents']['new_today_incidents'])
            for incident in pen_incidents:
                all_incidents.append([
                    pen_snap['pen_code'],
                    Paragraph(incident['title'], styles['Normal']),
                    incident['severity_display'],
                    incident['status_display'],
                    incident['incident_time'][:16].replace('T', ' '),
                    incident['reporter_name'],
                ])

        if all_incidents:
            incident_data = [incident_headers] + all_incidents
            incident_table = Table(
                incident_data,
                colWidths=[1.8 * cm, 6 * cm, 1.8 * cm, 1.8 * cm, 3 * cm, 2 * cm]
            )
            incident_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F8CBAD')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(incident_table)
        else:
            story.append(Paragraph('今日无异常事件', styles['Normal']))

        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="巡检日报_{target_date}.pdf"'
        return response

    @classmethod
    def export_daily_report(cls, target_date, format_type='excel', pen_id=None):
        """统一导出接口"""
        if format_type.lower() == 'pdf':
            return cls.export_daily_report_pdf(target_date, pen_id)
        else:
            return cls.export_daily_report_excel(target_date, pen_id)

    @classmethod
    def export_weekly_report_excel(cls, end_date, pen_id=None):
        """导出 Excel 格式的周报（汇总过去7天的数据）"""
        from datetime import timedelta
        start_date = end_date - timedelta(days=6)

        wb = Workbook()
        ws = wb.active
        ws.title = f'巡检周报_{start_date}_to_{end_date}'

        header_font = Font(bold=True, size=14, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        sub_header_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells('A1:H1')
        ws.cell(row=1, column=1, value=f'养殖场巡检周报').font = Font(bold=True, size=16)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:H2')
        ws.cell(row=2, column=1, value=f'统计周期: {start_date} 至 {end_date}').font = Font(size=11, italic=True)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A3:H3')
        ws.cell(row=3, column=1, value=f'导出时间: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}').font = Font(size=10, color='666666')
        ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')

        row = 5

        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row=row, column=1, value='每日汇总').font = sub_header_font
        row += 1

        daily_headers = ['日期', '巡检完成率', '巡检次数', '喂养次数', '清洁次数', '新增异常', '待处理事件', '整体状态']
        for col, header in enumerate(daily_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        daily_snapshots = []
        current_date = start_date
        while current_date <= end_date:
            snapshot = snapshot_service.calculate_daily_snapshot(current_date, pen_id)
            daily_snapshots.append(snapshot)
            current_date += timedelta(days=1)

        for snapshot in daily_snapshots:
            avg_rate = snapshot['average_inspection_completion_rate']
            total_inspections = sum(s['inspection']['inspection_count'] for s in snapshot['pen_snapshots'])
            total_feedings = sum(s['feeding']['feeding_count'] for s in snapshot['pen_snapshots'])
            total_cleanings = sum(s['cleaning']['cleaning_count'] for s in snapshot['pen_snapshots'])
            new_incidents = sum(s['incidents']['new_today_count'] for s in snapshot['pen_snapshots'])
            open_incidents = snapshot['total_open_incidents']

            if open_incidents > 0 or snapshot['total_abnormal_count'] > 0:
                status = '异常'
                status_color = cls.STATUS_COLORS['WARNING'].lstrip('#')
            elif avg_rate >= 100:
                status = '正常'
                status_color = cls.STATUS_COLORS['NORMAL'].lstrip('#')
            elif avg_rate >= 80:
                status = '部分完成'
                status_color = cls.STATUS_COLORS['PARTIAL'].lstrip('#')
            else:
                status = '需关注'
                status_color = cls.STATUS_COLORS['ATTENTION'].lstrip('#')

            row_data = [
                snapshot['date'],
                f'{avg_rate}%',
                total_inspections,
                total_feedings,
                total_cleanings,
                new_incidents,
                open_incidents,
                status
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = center_align
                cell.border = thin_border
                if col == 8:
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
            row += 1

        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row=row, column=1, value='栏区周汇总').font = sub_header_font
        row += 1

        pen_headers = ['栏区编号', '栏区名称', '平均完成率', '总巡检次数', '总喂养次数', '总清洁次数', '异常事件数', '整体状态']
        for col, header in enumerate(pen_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        pen_summary = {}
        for snapshot in daily_snapshots:
            for pen_snap in snapshot['pen_snapshots']:
                pen_id_val = pen_snap['pen_id']
                if pen_id_val not in pen_summary:
                    pen_summary[pen_id_val] = {
                        'pen_code': pen_snap['pen_code'],
                        'pen_name': pen_snap['pen_name'],
                        'completion_rates': [],
                        'total_inspections': 0,
                        'total_feedings': 0,
                        'total_cleanings': 0,
                        'total_incidents': 0
                    }
                pen_summary[pen_id_val]['completion_rates'].append(pen_snap['inspection']['completion_rate'])
                pen_summary[pen_id_val]['total_inspections'] += pen_snap['inspection']['inspection_count']
                pen_summary[pen_id_val]['total_feedings'] += pen_snap['feeding']['feeding_count']
                pen_summary[pen_id_val]['total_cleanings'] += pen_snap['cleaning']['cleaning_count']
                pen_summary[pen_id_val]['total_incidents'] += pen_snap['incidents']['new_today_count']

        for pen_data in pen_summary.values():
            avg_rate = round(sum(pen_data['completion_rates']) / len(pen_data['completion_rates']), 2) if pen_data['completion_rates'] else 0

            if pen_data['total_incidents'] > 0:
                status = '异常'
                status_color = cls.STATUS_COLORS['WARNING'].lstrip('#')
            elif avg_rate >= 100:
                status = '正常'
                status_color = cls.STATUS_COLORS['NORMAL'].lstrip('#')
            elif avg_rate >= 80:
                status = '部分完成'
                status_color = cls.STATUS_COLORS['PARTIAL'].lstrip('#')
            else:
                status = '需关注'
                status_color = cls.STATUS_COLORS['ATTENTION'].lstrip('#')

            row_data = [
                pen_data['pen_code'],
                pen_data['pen_name'],
                f'{avg_rate}%',
                pen_data['total_inspections'],
                pen_data['total_feedings'],
                pen_data['total_cleanings'],
                pen_data['total_incidents'],
                status
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = center_align if col != 2 else left_align
                cell.border = thin_border
                if col == 8:
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
            row += 1

        column_widths = [15, 20, 12, 12, 12, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="巡检周报_{start_date}_to_{end_date}.xlsx"'
        return response

    @classmethod
    def export_weekly_report_pdf(cls, end_date, pen_id=None):
        """导出 PDF 格式的周报（汇总过去7天的数据）"""
        from datetime import timedelta
        start_date = end_date - timedelta(days=6)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=10
        )
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            spaceBefore=10,
            spaceAfter=10
        )

        story = []

        story.append(Paragraph(f'养殖场巡检周报', title_style))
        story.append(Paragraph(f'统计周期: {start_date} 至 {end_date}', styles['Normal']))
        story.append(Paragraph(f'导出时间: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        story.append(Spacer(1, 0.5 * cm))

        daily_snapshots = []
        current_date = start_date
        while current_date <= end_date:
            snapshot = snapshot_service.calculate_daily_snapshot(current_date, pen_id)
            daily_snapshots.append(snapshot)
            current_date += timedelta(days=1)

        story.append(Paragraph('每日汇总', section_style))

        daily_data = [['日期', '巡检完成率', '巡检次数', '喂养次数', '清洁次数', '新增异常', '待处理事件', '整体状态']]

        for snapshot in daily_snapshots:
            avg_rate = snapshot['average_inspection_completion_rate']
            total_inspections = sum(s['inspection']['inspection_count'] for s in snapshot['pen_snapshots'])
            total_feedings = sum(s['feeding']['feeding_count'] for s in snapshot['pen_snapshots'])
            total_cleanings = sum(s['cleaning']['cleaning_count'] for s in snapshot['pen_snapshots'])
            new_incidents = sum(s['incidents']['new_today_count'] for s in snapshot['pen_snapshots'])
            open_incidents = snapshot['total_open_incidents']

            if open_incidents > 0 or snapshot['total_abnormal_count'] > 0:
                status = '异常'
                status_color = cls.STATUS_COLORS['WARNING']
            elif avg_rate >= 100:
                status = '正常'
                status_color = cls.STATUS_COLORS['NORMAL']
            elif avg_rate >= 80:
                status = '部分完成'
                status_color = cls.STATUS_COLORS['PARTIAL']
            else:
                status = '需关注'
                status_color = cls.STATUS_COLORS['ATTENTION']

            daily_data.append([
                snapshot['date'],
                f'{avg_rate}%',
                str(total_inspections),
                str(total_feedings),
                str(total_cleanings),
                str(new_incidents),
                str(open_incidents),
                status
            ])

        daily_table = Table(daily_data, colWidths=[2.5 * cm] * 8)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        for i in range(1, len(daily_data)):
            status_text = daily_data[i][7]
            if status_text == '异常':
                color = cls.STATUS_COLORS['WARNING']
            elif status_text == '正常':
                color = cls.STATUS_COLORS['NORMAL']
            elif status_text == '部分完成':
                color = cls.STATUS_COLORS['PARTIAL']
            else:
                color = cls.STATUS_COLORS['ATTENTION']
            style_cmds.append(('BACKGROUND', (7, i), (7, i), colors.HexColor(color)))

        daily_table.setStyle(TableStyle(style_cmds))
        story.append(daily_table)
        story.append(Spacer(1, 0.8 * cm))

        story.append(Paragraph('栏区周汇总', section_style))

        pen_summary = {}
        for snapshot in daily_snapshots:
            for pen_snap in snapshot['pen_snapshots']:
                pen_id_val = pen_snap['pen_id']
                if pen_id_val not in pen_summary:
                    pen_summary[pen_id_val] = {
                        'pen_code': pen_snap['pen_code'],
                        'pen_name': pen_snap['pen_name'],
                        'completion_rates': [],
                        'total_inspections': 0,
                        'total_feedings': 0,
                        'total_cleanings': 0,
                        'total_incidents': 0
                    }
                pen_summary[pen_id_val]['completion_rates'].append(pen_snap['inspection']['completion_rate'])
                pen_summary[pen_id_val]['total_inspections'] += pen_snap['inspection']['inspection_count']
                pen_summary[pen_id_val]['total_feedings'] += pen_snap['feeding']['feeding_count']
                pen_summary[pen_id_val]['total_cleanings'] += pen_snap['cleaning']['cleaning_count']
                pen_summary[pen_id_val]['total_incidents'] += pen_snap['incidents']['new_today_count']

        pen_data = [['栏区编号', '栏区名称', '平均完成率', '总巡检次数', '总喂养次数', '总清洁次数', '异常事件数', '整体状态']]

        for pen_data_item in pen_summary.values():
            avg_rate = round(sum(pen_data_item['completion_rates']) / len(pen_data_item['completion_rates']), 2) if pen_data_item['completion_rates'] else 0

            if pen_data_item['total_incidents'] > 0:
                status = '异常'
                status_color = cls.STATUS_COLORS['WARNING']
            elif avg_rate >= 100:
                status = '正常'
                status_color = cls.STATUS_COLORS['NORMAL']
            elif avg_rate >= 80:
                status = '部分完成'
                status_color = cls.STATUS_COLORS['PARTIAL']
            else:
                status = '需关注'
                status_color = cls.STATUS_COLORS['ATTENTION']

            pen_data.append([
                pen_data_item['pen_code'],
                pen_data_item['pen_name'],
                f'{avg_rate}%',
                str(pen_data_item['total_inspections']),
                str(pen_data_item['total_feedings']),
                str(pen_data_item['total_cleanings']),
                str(pen_data_item['total_incidents']),
                status
            ])

        pen_table = Table(pen_data, colWidths=[1.6 * cm, 3 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
        style_cmds_pen = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        for i in range(1, len(pen_data)):
            status_text = pen_data[i][7]
            if status_text == '异常':
                color = cls.STATUS_COLORS['WARNING']
            elif status_text == '正常':
                color = cls.STATUS_COLORS['NORMAL']
            elif status_text == '部分完成':
                color = cls.STATUS_COLORS['PARTIAL']
            else:
                color = cls.STATUS_COLORS['ATTENTION']
            style_cmds_pen.append(('BACKGROUND', (7, i), (7, i), colors.HexColor(color)))

        pen_table.setStyle(TableStyle(style_cmds_pen))
        story.append(pen_table)

        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="巡检周报_{start_date}_to_{end_date}.pdf"'
        return response

    @classmethod
    def export_weekly_report(cls, end_date, format_type='excel', pen_id=None):
        """周报统一导出接口"""
        if format_type.lower() == 'pdf':
            return cls.export_weekly_report_pdf(end_date, pen_id)
        else:
            return cls.export_weekly_report_excel(end_date, pen_id)


export_service = ExportService()
